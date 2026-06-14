"""샘플 test_cases.xlsx 생성 스크립트"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "http://localhost:5500"

test_cases = [
    # 화면명, URL, 액션타입, 셀렉터, 입력값, 기대결과
    ("로그인", f"{BASE_URL}/index.html", "navigate", "", "", ""),
    ("로그인_아이디입력", f"{BASE_URL}/index.html", "input", "#username", "admin", ""),
    ("로그인_비밀번호입력", f"{BASE_URL}/index.html", "input", "#password", "1234", ""),
    ("로그인_버튼클릭", f"{BASE_URL}/index.html", "click", "#btn-login", "", "URL contains dashboard"),
    ("목록_검색버튼", f"{BASE_URL}/index.html#/list", "navigate", "", "", ""),
    ("목록_검색실행", f"{BASE_URL}/index.html#/list", "click", "#btn-search", "", ".item-row count > 0"),
    ("목록_검색필터", f"{BASE_URL}/index.html#/list", "input", "#search-name", "홍", ""),
    ("목록_검색재실행", f"{BASE_URL}/index.html#/list", "click", "#btn-search", "", ".item-row count == 1"),
    ("등록_화면이동", f"{BASE_URL}/index.html#/register", "navigate", "", "", ""),
    ("등록_이름입력", f"{BASE_URL}/index.html#/register", "input", "#reg-name", "테스트유저", ""),
    ("등록_부서선택", f"{BASE_URL}/index.html#/register", "input", "#reg-dept", "개발팀", ""),
    ("등록_저장버튼", f"{BASE_URL}/index.html#/register", "click", "#btn-save", "", "text contains 등록 완료"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "테스트케이스"

headers = ["No", "화면명", "URL", "액션타입", "셀렉터", "입력값", "기대결과", "실제결과", "상태", "캡처경로"]
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 15
ws.column_dimensions["G"].width = 30
ws.column_dimensions["H"].width = 30
ws.column_dimensions["I"].width = 10
ws.column_dimensions["J"].width = 40

for i, (name, url, action, selector, value, expected) in enumerate(test_cases, 2):
    ws.cell(row=i, column=1, value=i-1)
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=url)
    ws.cell(row=i, column=4, value=action)
    ws.cell(row=i, column=5, value=selector)
    ws.cell(row=i, column=6, value=value)
    ws.cell(row=i, column=7, value=expected)

wb.save("test_cases.xlsx")
print(f"test_cases.xlsx 생성 완료 — {len(test_cases)}개 테스트케이스")
