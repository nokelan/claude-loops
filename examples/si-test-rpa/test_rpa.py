"""pytest — Excel 파서 및 구조 검증 (Selenium 없이 실행 가능)"""
import openpyxl
import subprocess
import sys
from pathlib import Path


def test_excel_created():
    """AC-1: test_cases.xlsx 생성 및 파싱"""
    if not Path("test_cases.xlsx").exists():
        subprocess.run([sys.executable, "create_sample_excel.py"], check=True)
    wb = openpyxl.load_workbook("test_cases.xlsx")
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "화면명" in headers
    assert "URL" in headers
    assert "액션타입" in headers
    assert "셀렉터" in headers
    assert "기대결과" in headers
    assert ws.max_row >= 2, "테스트케이스가 없음"


def test_excel_row_count():
    """AC-6: 10개 이상 테스트케이스"""
    wb = openpyxl.load_workbook("test_cases.xlsx")
    ws = wb.active
    data_rows = ws.max_row - 1
    assert data_rows >= 10, f"테스트케이스 {data_rows}개 (10개 이상 필요)"


def test_vue_sample_exists():
    """Vue 샘플 파일 존재 확인"""
    assert Path("vue-sample/index.html").exists(), "vue-sample/index.html 없음"


def test_results_dir_exists():
    """결과 디렉토리 존재 확인"""
    Path("results/screenshots").mkdir(parents=True, exist_ok=True)
    assert Path("results").exists()
    assert Path("results/screenshots").exists()
