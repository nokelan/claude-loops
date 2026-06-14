"""SI 테스트 자동화 RPA - 메인 실행기"""
import os
import sys
import re
import json
import time
import signal
import logging
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import InvalidSessionIdException, WebDriverException

# ── 설정 ──────────────────────────────────────────
INPUT_FILE = "test_cases.xlsx"
RESULTS_DIR = Path("results")
SCREENSHOTS_DIR = RESULTS_DIR / "screenshots"
LOG_FILE = RESULTS_DIR / "rpa.log"

RESULTS_DIR.mkdir(exist_ok=True)
SCREENSHOTS_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(color="276221", bold=True)
FAIL_FONT = Font(color="9C0006", bold=True)


# ── config.json 로드 ──────────────────────────────
def load_config() -> dict:
    config_path = Path("config.json")
    if config_path.exists():
        with open(config_path, encoding="utf-8") as f:
            cfg = json.load(f)
        log.info(f"config.json 로드: {cfg}")
        return cfg
    return {}


# ── 서버 자동 기동 ─────────────────────────────────
def start_server(cfg: dict):
    cmd = cfg.get("server_cmd")
    if not cmd:
        return None
    cwd = cfg.get("server_cwd", ".")
    log.info(f"서버 시작: {cmd} (cwd={cwd})")
    proc = subprocess.Popen(
        cmd, shell=True, cwd=cwd,
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    time.sleep(2)

    # health check
    base_url = cfg.get("base_url", "http://localhost:5500")
    try:
        import urllib.request
        urllib.request.urlopen(base_url, timeout=5)
        log.info(f"서버 응답 확인: {base_url}")
    except Exception as e:
        log.warning(f"서버 health check 실패: {e}")

    return proc


def stop_server(proc):
    if proc and proc.poll() is None:
        proc.terminate()
        log.info("서버 종료")


# ── 셀레니움 드라이버 ──────────────────────────────
def create_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--window-size=1280,800")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    return webdriver.Chrome(options=opts)


# ── 기대결과 검증 ─────────────────────────────────
def verify_expected(driver, expected: str) -> tuple[bool, str]:
    if not expected:
        return True, ""
    try:
        if expected.startswith("URL contains "):
            kw = expected.split("URL contains ", 1)[1]
            return kw in driver.current_url, driver.current_url

        if expected.startswith(".item-row count"):
            rows = driver.find_elements(By.CSS_SELECTOR, ".item-row")
            count = len(rows)
            m = re.search(r"count\s*([=><!]+)\s*(\d+)", expected)
            if m:
                op, val = m.group(1), int(m.group(2))
                ok = eval(f"{count} {op} {val}")
                return ok, f"count={count}"
            return count > 0, f"count={count}"

        if expected.startswith("text contains "):
            kw = expected.split("text contains ", 1)[1]
            body = driver.find_element(By.TAG_NAME, "body").text
            return kw in body, body[:80]

        if expected.startswith("text =="):
            m = re.search(r"text\s*==\s*(.+)", expected)
            if m:
                parts = m.group(1).strip().split(" ", 1)
                sel, val = (parts[0], parts[1]) if len(parts) == 2 else ("body", parts[0])
                el = driver.find_element(By.CSS_SELECTOR, sel)
                actual = el.text.strip()
                return actual == val.strip(), actual

        return True, "(검증 없음)"
    except Exception as e:
        return False, f"검증 오류: {e}"


# ── navigate 검증 ──────────────────────────────────
def navigate_and_verify(driver, url: str, wait_selector: str = None):
    """navigate 후 페이지 실제 로드 확인. wait_selector로 element 존재 체크."""
    driver.get(url)
    time.sleep(0.5)
    if wait_selector:
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, wait_selector))
            )
        except Exception:
            log.warning(f"navigate 후 element 없음: {wait_selector} on {url}")
            return False
    return True


# ── 단일 케이스 실행 ──────────────────────────────
def run_case(driver, row: dict, ts: str, retry_create_driver=None) -> dict:
    action = row.get("액션타입", "")
    url = row.get("URL", "")
    selector = row.get("셀렉터", "") or ""
    value = row.get("입력값", "") or ""
    expected = row.get("기대결과", "") or ""
    name = row.get("화면명", "")

    actual = ""
    status = "FAIL"

    def _execute():
        nonlocal actual, status
        wait = WebDriverWait(driver, 8)

        if action == "navigate":
            ok = navigate_and_verify(driver, url, selector or None)
            status = "PASS" if ok else "FAIL"
            if not ok:
                actual = f"navigate 실패: {url}"

        elif action == "input":
            if driver.current_url.split("#")[0] != url.split("#")[0]:
                driver.get(url)
                time.sleep(0.5)
            el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
            if el.tag_name.lower() == "select":
                Select(el).select_by_visible_text(value)
            else:
                el.clear()
                el.send_keys(value)
            status = "PASS"

        elif action == "click":
            if driver.current_url.split("#")[0] != url.split("#")[0]:
                driver.get(url)
                time.sleep(0.5)
            el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
            el.click()
            time.sleep(0.8)
            ok, actual = verify_expected(driver, expected)
            status = "PASS" if ok else "FAIL"

        elif action == "check_text":
            if driver.current_url.split("#")[0] != url.split("#")[0]:
                driver.get(url)
                time.sleep(0.5)
            el = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
            actual = el.text.strip()
            m = re.search(r"text\s*==\s*(.+)", expected)
            if m:
                status = "PASS" if actual == m.group(1).strip() else "FAIL"
            else:
                status = "PASS"

    try:
        _execute()
    except (InvalidSessionIdException, WebDriverException) as e:
        msg = str(e)
        if "invalid session" in msg.lower() and retry_create_driver:
            log.warning(f"[{name}] 세션 크래시 감지 → 드라이버 재시작 후 재시도")
            new_driver = retry_create_driver()
            # 드라이버 교체 후 재시도 (caller가 처리)
            actual = "SESSION_CRASH"
            status = "RETRY"
        else:
            actual = msg[:100]
            status = "FAIL"
            log.error(f"[{name}] 오류: {e}")
    except Exception as e:
        actual = str(e)[:100]
        status = "FAIL"
        log.error(f"[{name}] 오류: {e}")

    # 캡처
    cap_path = ""
    try:
        safe_name = re.sub(r"[^\w가-힣]", "_", name)
        cap_file = SCREENSHOTS_DIR / f"{ts}_{safe_name}.png"
        driver.save_screenshot(str(cap_file))
        cap_path = str(cap_file)
    except Exception as e:
        log.warning(f"캡처 실패: {e}")

    log.info(f"[{name}] {status} | 실제: {actual or '-'} | 캡처: {cap_path}")
    return {"실제결과": actual, "상태": status, "캡처경로": cap_path}


# ── 메인 ─────────────────────────────────────────
def main(input_file: str = INPUT_FILE):
    cfg = load_config()

    if not Path(input_file).exists():
        log.error(f"{input_file} 없음. create_sample_excel.py를 먼저 실행하세요.")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_file = RESULTS_DIR / f"{ts}_report.xlsx"

    server_proc = start_server(cfg)
    driver = create_driver()
    log.info(f"=== SI 테스트 RPA 시작 ({ts}) ===")

    pass_count = fail_count = 0

    actual_col = headers.index("실제결과") + 1
    status_col = headers.index("상태") + 1
    cap_col = headers.index("캡처경로") + 1

    try:
        for row_idx in range(2, ws.max_row + 1):
            row = {headers[c]: ws.cell(row_idx, c + 1).value for c in range(len(headers))}
            if not row.get("URL"):
                continue

            result = run_case(driver, row, ts, retry_create_driver=create_driver)

            # 세션 크래시 시 드라이버 교체 후 재시도
            if result["상태"] == "RETRY":
                log.info(f"[{row.get('화면명')}] 드라이버 재시작 후 재시도")
                try:
                    driver.quit()
                except Exception:
                    pass
                driver = create_driver()
                result = run_case(driver, row, ts)

            ws.cell(row_idx, actual_col, result["실제결과"])
            status_cell = ws.cell(row_idx, status_col, result["상태"])
            ws.cell(row_idx, cap_col, result["캡처경로"])

            if result["상태"] == "PASS":
                status_cell.fill = PASS_FILL
                status_cell.font = PASS_FONT
                pass_count += 1
            else:
                status_cell.fill = FAIL_FILL
                status_cell.font = FAIL_FONT
                fail_count += 1

    finally:
        try:
            driver.quit()
        except Exception:
            pass
        stop_server(server_proc)

    wb.save(result_file)
    log.info(f"=== 완료: PASS {pass_count} / FAIL {fail_count} ===")
    log.info(f"결과 리포트 저장: {result_file}")
    print(f"\n결과 리포트 Excel: {result_file}")
    return pass_count, fail_count


if __name__ == "__main__":
    f = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    main(f)
